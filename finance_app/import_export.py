"""Import and export helpers for CSV, Excel and PDF files."""

import csv
from pathlib import Path
from tkinter import filedialog, messagebox

try:  # optional dependencies
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover - optional import
    openpyxl = None

try:
    from PyPDF2 import PdfReader  # type: ignore
except Exception:  # pragma: no cover - optional import
    PdfReader = None

from .transactions import TransactionManager, Transaction

class ImportExport:
    """Handle CSV, Excel and PDF import/export with simple dialogs."""

    def __init__(self, manager: TransactionManager, master=None):
        self.manager = manager
        self.master = master

    def import_csv_dialog(self):
        file_path = filedialog.askopenfilename(
            parent=self.master,
            title='Import CSV',
            filetypes=[('CSV files', '*.csv')]
        )
        if file_path:
            try:
                self.manager.import_csv(file_path)
                messagebox.showinfo('Import', 'Import successful')
            except Exception as e:
                messagebox.showerror('Import failed', str(e))

    def export_csv_dialog(self):
        file_path = filedialog.asksaveasfilename(
            parent=self.master,
            title='Export CSV',
            defaultextension='.csv',
            filetypes=[('CSV files', '*.csv')]
        )
        if file_path:
            try:
                self.manager.export_csv(file_path)
                messagebox.showinfo('Export', 'Export successful')
            except Exception as e:
                messagebox.showerror('Export failed', str(e))

    def export_summary_dialog(self):
        """Export a basic summary by category to a CSV file."""
        file_path = filedialog.asksaveasfilename(
            parent=self.master,
            title='Export Summary',
            defaultextension='.csv',
            filetypes=[('CSV files', '*.csv')]
        )
        if file_path:
            try:
                summary = self.manager.summary_by_category()
                with open(file_path, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['category', 'spent'])
                    for cat, spent in summary.items():
                        writer.writerow([cat, spent])
                messagebox.showinfo('Export', 'Export successful')
            except Exception as e:
                messagebox.showerror('Export failed', str(e))

    # ------------------------------------------------------------------
    # Excel helpers
    def import_excel_dialog(self):
        """Import transactions from an Excel file if openpyxl is available."""
        file_path = filedialog.askopenfilename(
            parent=self.master,
            title='Import Excel',
            filetypes=[('Excel files', '*.xlsx')],
        )
        if file_path:
            if not openpyxl:
                messagebox.showerror('Import failed', 'openpyxl not installed')
                return
            try:
                self.import_excel(Path(file_path))
                messagebox.showinfo('Import', 'Import successful')
            except Exception as e:
                messagebox.showerror('Import failed', str(e))

    def export_excel_dialog(self):
        """Export transactions to an Excel file if openpyxl is available."""
        file_path = filedialog.asksaveasfilename(
            parent=self.master,
            title='Export Excel',
            defaultextension='.xlsx',
            filetypes=[('Excel files', '*.xlsx')],
        )
        if file_path:
            if not openpyxl:
                messagebox.showerror('Export failed', 'openpyxl not installed')
                return
            try:
                self.export_excel(Path(file_path))
                messagebox.showinfo('Export', 'Export successful')
            except Exception as e:
                messagebox.showerror('Export failed', str(e))

    def import_excel(self, path: Path):
        """Load transactions from an Excel worksheet."""
        wb = openpyxl.load_workbook(path)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            date, desc, amount, category = row
            self.manager.add_transaction(
                Transaction(date=str(date), description=str(desc), amount=float(amount), category=str(category))
            )

    def export_excel(self, path: Path):
        """Save transactions to an Excel worksheet."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(['date', 'description', 'amount', 'category'])
        for t in self.manager.transactions:
            sheet.append([t.date, t.description, t.amount, t.category])
        wb.save(path)

    # ------------------------------------------------------------------
    # PDF helpers
    def import_pdf_dialog(self):
        """Import transactions from a PDF file if PyPDF2 is available."""
        file_path = filedialog.askopenfilename(
            parent=self.master,
            title='Import PDF',
            filetypes=[('PDF files', '*.pdf')],
        )
        if file_path:
            if not PdfReader:
                messagebox.showerror('Import failed', 'PyPDF2 not installed')
                return
            try:
                self.import_pdf(Path(file_path))
                messagebox.showinfo('Import', 'Import successful')
            except Exception as e:
                messagebox.showerror('Import failed', str(e))

    def import_pdf(self, path: Path):
        """Very simple PDF table reader expecting comma-separated lines."""
        reader = PdfReader(str(path))
        text = "\n".join(page.extract_text() or '' for page in reader.pages)
        for line in text.splitlines():
            parts = [p.strip() for p in line.split(',')]
            if len(parts) != 4:
                continue
            date, desc, amount, category = parts
            try:
                amount = float(amount)
            except ValueError:
                continue
            self.manager.add_transaction(
                Transaction(date=date, description=desc, amount=amount, category=category)
            )
